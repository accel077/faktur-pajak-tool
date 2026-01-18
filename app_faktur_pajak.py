import streamlit as st
import pdfplumber
from PyPDF2 import PdfMerger
import re
import io
import zipfile
import pandas as pd
from openpyxl import load_workbook

# --- KONFIGURASI HALAMAN ---
st.set_page_config(page_title="Faktur Pajak Tool Pro", layout="wide")

# --- FUNGSI LOGIKA ---
def extract_referensi(pdf_file):
    try:
        with pdfplumber.open(pdf_file) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    match = re.search(r'PJ\d+', text)
                    if match: return match.group(0)
        return None
    except: return None

def get_color_mapping(excel_file):
    color_map = {}
    try:
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active
        ref_col_idx = None
        for cell in ws[1]:
            if cell.value == "Referensi":
                ref_col_idx = cell.column
                break
        if not ref_col_idx: return None
        for row in ws.iter_rows(min_row=2):
            cell = row[ref_col_idx-1]
            ref_val = str(cell.value).strip() if cell.value else None
            if ref_val and ref_val.startswith("PJ"):
                fill = cell.fill
                color_hex = fill.start_color.index if fill and hasattr(fill.start_color, 'index') else None
                color_map[ref_val] = str(color_hex)
        return color_map
    except: return None

# --- FUNGSI RESET ---
def reset_tab(prefix):
    """Menghapus data di session state dan mengubah key file uploader"""
    st.session_state[f"{prefix}_data"] = None
    st.session_state[f"{prefix}_results"] = None
    st.session_state[f"{prefix}_uploader_key"] += 1

# --- INISIALISASI SESSION STATE ---
for prefix in ["ren", "cls", "mrg"]:
    if f"{prefix}_data" not in st.session_state: st.session_state[f"{prefix}_data"] = None
    if f"{prefix}_results" not in st.session_state: st.session_state[f"{prefix}_results"] = None
    if f"{prefix}_uploader_key" not in st.session_state: st.session_state[f"{prefix}_uploader_key"] = 0

# --- SIDEBAR ---
with st.sidebar:
    st.header("⚙️ Database")
    excel_db = st.file_uploader("Upload Excel Warna", type=["xlsx"])
    color_map = get_color_mapping(excel_db) if excel_db else None
    if color_map: st.success("✅ Database Dimuat!")
    st.divider()
    st.warning("Gunakan tombol 'Clear All' untuk mengosongkan antrean file.")

st.title("📑 Faktur Pajak Tool Pro")
tab1, tab2, tab3 = st.tabs(["🔄 Rename", "📁 Klasifikasi ZIP", "🔗 Merge PDF"])

# --- TAB 1: RENAME ---
with tab1:
    col_head, col_btn = st.columns([0.8, 0.2])
    with col_head: st.subheader("Rename dengan Live Preview")
    with col_btn: 
        if st.button("🗑️ Clear All", key="btn_clear_ren"): reset_tab("ren")

    files_ren = st.file_uploader("Upload PDF", type="pdf", accept_multiple_files=True, 
                                 key=f"ren_up_{st.session_state.ren_uploader_key}")
    
    if st.button("Mulai Proses Rename", type="primary"):
        if files_ren:
            results, ren_buffer = [], io.BytesIO()
            p_bar = st.progress(0)
            with zipfile.ZipFile(ren_buffer, "a", zipfile.ZIP_DEFLATED) as zip_f:
                for i, f in enumerate(files_ren):
                    ref = extract_referensi(f)
                    name_only = f.name.replace(".pdf", "").replace(".PDF", "")
                    parts = name_only.split('-')
                    suffix = "-".join(parts[-3:]) if len(parts) >= 3 else name_only
                    new_name = f"{ref} {suffix}.pdf" if ref else f.name
                    if ref: zip_f.writestr(new_name, f.getvalue())
                    results.append({"Status": "✅ Berhasil" if ref else "❌ Gagal", "Nama Asli": f.name, "Nama Baru": new_name})
                    p_bar.progress((i + 1) / len(files_ren))
            st.session_state.ren_data = ren_buffer.getvalue()
            st.session_state.ren_results = results

    if st.session_state.ren_results:
        st.table(pd.DataFrame(st.session_state.ren_results))
        st.download_button("⬇️ Download ZIP Rename", st.session_state.ren_data, "rename_faktur.zip", use_container_width=True)

# --- TAB 2: KLASIFIKASI ---
with tab2:
    col_head2, col_btn2 = st.columns([0.8, 0.2])
    with col_head2: st.subheader("Klasifikasi ke Folder")
    with col_btn2:
        if st.button("🗑️ Clear All", key="btn_clear_cls"): reset_tab("cls")

    if not color_map: st.warning("Upload Excel di sidebar dahulu.")
    else:
        files_cls = st.file_uploader("Upload PDF", type="pdf", accept_multiple_files=True, 
                                     key=f"cls_up_{st.session_state.cls_uploader_key}")
        
        if st.button("Mulai Klasifikasi", type="primary"):
            if files_cls:
                results_cls, cls_buffer = [], io.BytesIO()
                p_bar_cls = st.progress(0)
                with zipfile.ZipFile(cls_buffer, "a", zipfile.ZIP_DEFLATED) as zip_c:
                    for i, f in enumerate(files_cls):
                        ref = extract_referensi(f)
                        folder = color_map.get(ref, "Tidak_Ada_Di_Excel")
                        zip_c.writestr(f"{folder}/{f.name}", f.getvalue())
                        results_cls.append({"File": f.name, "Folder": folder, "Status": "✅ OK" if ref else "⚠️ No Ref"})
                        p_bar_cls.progress((i + 1) / len(files_cls))
                st.session_state.cls_data = cls_buffer.getvalue()
                st.session_state.cls_results = results_cls

        if st.session_state.cls_results:
            st.table(pd.DataFrame(st.session_state.cls_results))
            st.download_button("⬇️ Download ZIP Klasifikasi", st.session_state.cls_data, "klasifikasi.zip", use_container_width=True)

# --- TAB 3: MERGE ---
with tab3:
    col_head3, col_btn3 = st.columns([0.8, 0.2])
    with col_head3: st.subheader("Gabungkan PDF (A-Z)")
    with col_btn3:
        if st.button("🗑️ Clear All", key="btn_clear_mrg"): reset_tab("mrg")

    files_mrg = st.file_uploader("Upload PDF", type="pdf", accept_multiple_files=True, 
                                 key=f"mrg_up_{st.session_state.mrg_uploader_key}")
    
    if st.button("Proses Gabung", type="primary"):
        if len(files_mrg) >= 2:
            merger = PdfMerger()
            sorted_files = sorted(files_mrg, key=lambda x: x.name)
            p_bar_m = st.progress(0)
            for i, f in enumerate(sorted_files):
                merger.append(f)
                p_bar_m.progress((i + 1) / len(sorted_files))
            m_out = io.BytesIO()
            merger.write(m_out)
            st.session_state.mrg_data = m_out.getvalue()
            st.success("✅ Penggabungan Selesai!")

    if st.session_state.mrg_data:
        st.download_button("⬇️ Download PDF Gabungan", st.session_state.mrg_data, "hasil_merge.pdf", use_container_width=True)
